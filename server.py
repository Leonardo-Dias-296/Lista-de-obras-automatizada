from flask import Flask, request, send_file, jsonify, send_from_directory
import openpyxl
import json
import math
import io
import os
from datetime import datetime
from fpdf import FPDF

app = Flask(__name__, static_folder='public', static_url_path='/static')

# ─── Diretório base (garante caminhos corretos em qualquer cwd) ───────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ─── Mapeamento (O antigo foi deletado para escaneamento autônomo na Fase 8) ───
import unicodedata

def normalize_name(name):
    if not name:
        return ""
    name = str(name).strip().lower()
    name = "".join(
        c for c in unicodedata.normalize('NFD', name)
        if unicodedata.category(c) != 'Mn'
    )
    return "".join(c for c in name if c.isalnum())

# ─── Regras de cálculo por módulo ─────────────────────────────────────────────
def calcular_terra(m):
    valor = math.floor(m / 2) + 2
    if valor % 2 != 0:
        valor -= 1
    if valor >= m:
        valor = m - 1
        if valor % 2 != 0:
            valor -= 1
    return valor

REGRAS_MODULO = {
    "Perfil":                               lambda m: m,
    "Emenda Barra":                         lambda m: math.ceil((m + 2) / 2) * 2,
    "Emenda Placa":                         lambda m: (2 * m) + 8,
    "Canto / Final":                        lambda m: math.ceil((m / 2) * 4 + 4),
    "Terra":                                calcular_terra,
    "Conjunto PE em L PRISIONEIRO":         lambda m: (2 * m) + 8,
    "Placa propaganda P":                   lambda m: 1 if m < 30 else 0,
    "Placa propaganda G":                   lambda m: 1 if m >= 30 else 0,
}

# ─── Carregar banco de dados ──────────────────────────────────────────────────
def load_db():
    with open(os.path.join(BASE_DIR, 'database.json'), 'r', encoding='utf-8') as f:
        return json.load(f)

# ─── Calcular todas as quantidades ───────────────────────────────────────────
def calcular_quantidades(modulos_lista, inversores, cidade_nome, db):
    quantidades = {}
    
    total_modulos = sum(int(m.get('qtd', 0)) for m in modulos_lista) if isinstance(modulos_lista, list) else int(modulos_lista)

    # 1. Regras por módulo
    for nome, regra in REGRAS_MODULO.items():
        qtd = regra(total_modulos)
        if qtd > 0:
            quantidades[nome] = qtd

    # 2. Regras por inversor
    for inv in inversores:
        inv_nome = inv.get('nome', '').strip()
        inv_qtd = int(inv.get('qtd', 1))
        if inv_nome in db.get('inversores', {}):
            for item in db['inversores'][inv_nome].get('materiais', []):
                nome = item.get('nome')
                cod  = item.get('codigo')
                qtd  = item.get('qtd', 1) * inv_qtd
                if cod:
                    # usa "CODIGO_SEPARATOR_NOME" na key para n perder dados
                    quantidades[f"[{cod}] {nome}"] = quantidades.get(f"[{cod}] {nome}", 0) + qtd
                else:
                    quantidades[nome] = quantidades.get(nome, 0) + qtd

    # 3. Itens fixos
    for item in db.get('itensFixos', []):
        nome = item['nome']
        qtd  = item['qtd']
        quantidades[nome] = quantidades.get(nome, 0) + qtd

    # 4. Exclusões por cidade
    if cidade_nome in db.get('cidades', {}):
        for item in db['cidades'][cidade_nome].get('naoadicionar', []):
            nome = item['nome']
            if nome in quantidades:
                del quantidades[nome]

    # 5. Regra especial para o Conector MC4
    total_micro = 0
    for inv in inversores:
        if "micro" in inv.get('nome', '').lower():
            total_micro += int(inv.get('qtd', 1))

    if total_micro > 0:
        qtd_mc4 = (8 * total_micro) + 8
    else:
        if total_modulos <= 20:
            qtd_mc4 = total_modulos - 2
        elif total_modulos <= 50:
            qtd_mc4 = total_modulos - 10
        elif total_modulos <= 100:
            qtd_mc4 = total_modulos - 20
        else:
            qtd_mc4 = total_modulos - 30
        if qtd_mc4 < 0:
            qtd_mc4 = 0

    if qtd_mc4 > 0:
        quantidades['Conector MC4'] = qtd_mc4

    return quantidades

# ─── Log de Histórico para Relatórios ─────────────────────────────────────────
import hashlib

def anonimizar_nome(nome):
    """Cria hash do nome para anonimização LGPD."""
    if not nome:
        return "ANONIMO"
    hash_obj = hashlib.md5(f"ssm_{nome}".encode())
    return f"Cliente_{hash_obj.hexdigest()[:8].upper()}"

def limpar_historico_antigo():
    """Remove registros com mais de 6 meses (política de retenção LGPD)."""
    try:
        caminho = os.path.join(BASE_DIR, 'history.json')
        if not os.path.exists(caminho):
            return
        
        with open(caminho, 'r', encoding='utf-8') as f:
            historico = json.load(f)
        
        data_limite = datetime.now().replace(month=datetime.now().month - 6 if datetime.now().month > 6 else datetime.now().month + 6)
        if datetime.now().month <= 6:
            data_limite = data_limite.replace(year=datetime.now().year - 1)
        
        historico_filtrado = []
        for entry in historico:
            ts = entry.get('ts_geracao', '')
            if ts:
                try:
                    data_entry = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
                    if data_entry >= data_limite:
                        historico_filtrado.append(entry)
                except:
                    historico_filtrado.append(entry)
            else:
                historico_filtrado.append(entry)
        
        with open(caminho, 'w', encoding='utf-8') as f:
            json.dump(historico_filtrado, f, ensure_ascii=False, indent=2)
        
        removidos = len(historico) - len(historico_filtrado)
        if removidos > 0:
            print(f"[LGPD] {removidos} registros antigos removidos do histórico")
    except Exception as e:
        print(f"[ERRO] Falha ao limpar histórico: {e}")

def log_history(dados):
    """Salva os dados do orçamento no histórico para relatórios (LGPD compliant)."""
    try:
        limpar_historico_antigo()
        
        caminho = os.path.join(BASE_DIR, 'history.json')
        historico = []
        if os.path.exists(caminho):
            with open(caminho, 'r', encoding='utf-8') as f:
                historico = json.load(f)
        
        # Se for um lote, dados é uma lista; se for unitário, um dict
        if isinstance(dados, list):
            for item in dados:
                item_anon = item.copy()
                item_anon['cliente'] = anonimizar_nome(item.get('cliente', ''))
                item_anon['ts_geracao'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                historico.append(item_anon)
        else:
            dados_anon = dados.copy()
            dados_anon['cliente'] = anonimizar_nome(dados.get('cliente', ''))
            dados_anon['ts_geracao'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            historico.append(dados_anon)
            
        with open(caminho, 'w', encoding='utf-8') as f:
            json.dump(historico, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[ERRO] Falha ao logar histórico: {e}")

# ─── Rotas ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_file('index.html')

@app.route('/privacidade')
def privacidade():
    return send_file(os.path.join(BASE_DIR, 'public', 'privacidade.html'))

@app.route('/api/config')
def config():
    db = load_db()
    return jsonify(db)

@app.route('/api/gerar', methods=['POST'])
def gerar():
    try:
        data = request.get_json()

        cliente     = data.get('cliente', '').strip()
        data_obra   = data.get('data', '').strip()
        equipe      = data.get('equipe', '').strip()
        cod_obra    = data.get('cod_obra', '').strip()
        modulos     = data.get('modulos', [])
        inversores  = data.get('inversores', [])
        extras      = data.get('extras', [])
        cidade      = data.get('cidade', '').strip()
        estrutura   = data.get('estrutura', '').strip()

        db = load_db()
        quantidades = calcular_quantidades(modulos, inversores, cidade, db)

        # Abrir modelo sem alterar o original
        wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'modelo.xlsx'))
        ws = wb.active

        # Preencher cabeçalho (Linhas 2-7 estão mescladas A:C)
        def preencher_ajustado(row, label, valor):
            orig = ws.cell(row=row, column=1).value or ""
            # Se o label já existir no texto original, preservamos ele e adicionamos o valor
            if label.strip() in orig:
                ws.cell(row=row, column=1).value = f"{label}{valor}"
            else:
                ws.cell(row=row, column=1).value = f"{label}: {valor}"

        inversores_str = ", ".join([f"{i.get('qtd', 1)}x {i.get('nome', '')}" for i in inversores])
        modulos_str = ", ".join([f"{m.get('qtd', 1)}x {m.get('potencia', '')}W" for m in modulos]) if modulos else "Nenhum"

        preencher_ajustado(2, "Data: ", data_obra)
        preencher_ajustado(3, "Cliente: ", cliente)
        preencher_ajustado(4, "Equipe: ", equipe)
        preencher_ajustado(5, "Inversor/Marca: ", inversores_str)
        preencher_ajustado(6, "Paineis (Qtde/Potencia): ", modulos_str)
        preencher_ajustado(7, "Cód obra: ", cod_obra)
        if estrutura:
            preencher_ajustado(8, "Estrutura: ", estrutura)

        # Criar map das linhas originais da planilha pela Coluna 1
        row_map = {}
        for r in range(9, 56):
            c = ws.cell(row=r, column=1).value
            if c is not None:
                row_map[str(c).strip()] = r
                
        # Dicionário Nome Normalizado -> Código
        nome_para_codigo = {normalize_name(p['nome']): str(p['codigo']) for p in db.get('produtos', [])}

        extra_row = 57 # Começa as escritas extras
        for ext in extras:
            nome_ext = ext.get('nome', '').strip()
            qtd_ext = ext.get('qtd', '')
            cod = ext.get('codigo')
            if not nome_ext and not cod: continue
            
            nome_norm = normalize_name(nome_ext)
            # Tenta resolver código explícito ou implícito
            if cod and str(cod) in row_map:
                ws.cell(row=row_map[str(cod)], column=3).value = qtd_ext
            elif nome_norm in nome_para_codigo and nome_para_codigo[nome_norm] in row_map:
                ws.cell(row=row_map[nome_para_codigo[nome_norm]], column=3).value = qtd_ext
            else:
                ws.cell(row=extra_row, column=2).value = nome_ext
                ws.cell(row=extra_row, column=3).value = qtd_ext
                extra_row += 1

        # Disjuntor CA e CC (linha 28 e 29)
        disj_ca_lista = []
        disj_cc_lista = []
        for inv in inversores:
            i_nome = inv.get('nome', '').strip()
            if i_nome in db.get('inversores', {}):
                dca = db['inversores'][i_nome].get('disjuntor_ca', '')
                dcc = db['inversores'][i_nome].get('disjuntor_cc', '')
                if dca: disj_ca_lista.append(f"{inv.get('qtd')}x {dca}" if inv.get('qtd', 1) > 1 else dca)
                if dcc: disj_cc_lista.append(f"{inv.get('qtd')}x {dcc}" if inv.get('qtd', 1) > 1 else dcc)
                
        if disj_ca_lista:
            ws.cell(row=28, column=3).value = " / ".join(disj_ca_lista)
        if disj_cc_lista:
            ws.cell(row=29, column=3).value = " / ".join(disj_cc_lista)

        # Preencher quantidades de Materiais (Estrutura e Inversores)
        for key, qtd in quantidades.items():
            nome = key
            cod = None
            if key.startswith('[') and '] ' in key:
                cod = key.split('] ')[0][1:]
                nome = key.split('] ')[1]
            
            nome_norm = normalize_name(nome)
            if cod and str(cod) in row_map:
                ws.cell(row=row_map[str(cod)], column=3).value = qtd
            elif nome_norm in nome_para_codigo and nome_para_codigo[nome_norm] in row_map:
                ws.cell(row=row_map[nome_para_codigo[nome_norm]], column=3).value = qtd
            else:
                # É um item não listado nativo, envia pra base extra
                ws.cell(row=extra_row, column=2).value = nome
                ws.cell(row=extra_row, column=3).value = qtd
                extra_row += 1

        # ── Salvar em disco na pasta saidas/ ─────────────────────────────────
        pasta_saidas = os.path.join(BASE_DIR, 'saidas')
        os.makedirs(pasta_saidas, exist_ok=True)

        nome_seguro   = "".join(c for c in cliente if c.isalnum() or c in ' _-').strip().replace(' ', '_')
        filename      = f"SSM_{nome_seguro}_{data_obra}.xlsx"
        caminho_disco = os.path.join(pasta_saidas, filename)

        wb.save(caminho_disco)
        print(f"[OK] Arquivo salvo em: {caminho_disco}")

        # Logar histórico
        log_history(data)

        # ── Enviar como download para o navegador ─────────────────────────────
        return send_file(
            caminho_disco,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

@app.route('/api/preview', methods=['POST'])
def preview():
    """Retorna as quantidades calculadas para prévia no site."""
    try:
        data       = request.get_json()
        modulos    = data.get('modulos', [])
        inversores = data.get('inversores', [])
        cidade     = data.get('cidade', '').strip()

        db = load_db()
        quantidades = calcular_quantidades(modulos, inversores, cidade, db)
        return jsonify(quantidades)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/arquivos')
def listar_arquivos():
    """Lista os arquivos gerados na pasta saidas/."""
    SAIDAS_DIR = os.path.join(BASE_DIR, 'saidas')
    if not os.path.exists(SAIDAS_DIR): os.makedirs(SAIDAS_DIR)
    files = []
    for f in os.listdir(SAIDAS_DIR):
        if f.endswith('.xlsx') or f.endswith('.pdf'):
            path = os.path.join(SAIDAS_DIR, f)
            files.append({
                "nome": f,
                "data": datetime.fromtimestamp(os.path.getctime(path)).strftime("%d/%m/%Y %H:%M")
            })
    return jsonify(sorted(files, key=lambda x: x['data'], reverse=True))

@app.route('/api/arquivos/<path:filename>', methods=['DELETE'])
def deletar_arquivo(filename):
    """Deleta fisicamente um arquivo da pasta saidas/."""
    try:
        pasta = os.path.join(BASE_DIR, 'saidas')
        caminho = os.path.join(pasta, filename)
        caminho_real = os.path.realpath(caminho)
        pasta_real = os.path.realpath(pasta)
        if not caminho_real.startswith(pasta_real):
            return jsonify({'erro': 'Acesso negado'}), 403
        if os.path.exists(caminho):
            os.remove(caminho)
            return jsonify({'status': 'ok'})
        return jsonify({'erro': 'Arquivo não encontrado'}), 404
    except Exception as e:
        return jsonify({'erro': 'Erro ao processar solicitação'}), 500

@app.route('/api/stats')
def get_stats():
    HISTORY_FILE = 'history.json'
    history = []
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            history = json.load(f)
    
    total = len(history)
    inv_counts = {}
    city_counts = {}
    total_modulos = 0
    
    for entry in history:
        inv = entry.get('inversor', 'Desconhecido')
        city = entry.get('cidade', 'Desconhecida')
        mods = entry.get('modulos', 0)
        
        inv_counts[inv] = inv_counts.get(inv, 0) + 1
        city_counts[city] = city_counts.get(city, 0) + 1
        total_modulos += int(mods) if isinstance(mods, (int, str)) else 0
        
    # Sort top items
    top_inv = sorted(inv_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    top_city = sorted(city_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    
    return jsonify({
        "total_orcamentos": total,
        "total_modulos": total_modulos,
        "top_inversores": dict(top_inv),
        "top_cidades": dict(top_city)
    })

@app.route('/api/gerar_pdf', methods=['POST'])
def gerar_pdf():
    dados = request.json
    cliente = dados.get('cliente', 'Cliente')
    equipe = dados.get('equipe', '—')
    data_obra = dados.get('data', datetime.now().strftime("%Y-%m-%d"))
    modulos = dados.get('modulos', [])
    modulos_str = ", ".join([f"{m.get('qtd', 1)}x {m.get('potencia', '')}W" for m in modulos]) if modulos else "Nenhum"
    inversores = dados.get('inversores', [])
    inversores_str = ", ".join([f"{i.get('qtd', 1)}x {i.get('nome', '')}" for i in inversores]) if inversores else "—"
    materiais = dados.get('materiais', {}) # Formato: {nome: qtd}
    
    pdf = FPDF()
    pdf.add_page()
    
    # Header
    pdf.set_fill_color(15, 23, 42) # Dark blue
    pdf.rect(0, 0, 210, 40, 'F')
    
    pdf.set_font("Arial", 'B', 20)
    pdf.set_text_color(251, 191, 36) # Sun yellow
    pdf.cell(0, 15, "SSM SOLAR - ROMANEIO DE CARGA", 0, 1, 'C')
    
    pdf.set_font("Arial", '', 10)
    pdf.set_text_color(200, 200, 200)
    pdf.cell(0, 5, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
    pdf.ln(15)
    
    # Info Section
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 10, f"DADOS DA OBRA", "B", 1)
    pdf.set_font("Arial", '', 11)
    
    col1 = 40
    pdf.cell(col1, 8, "Cliente:", 0, 0); pdf.cell(0, 8, str(cliente), 0, 1)
    pdf.cell(col1, 8, "Data Obra:", 0, 0); pdf.cell(0, 8, str(data_obra), 0, 1)
    pdf.cell(col1, 8, "Equipe:", 0, 0); pdf.cell(0, 8, str(equipe), 0, 1)
    pdf.cell(col1, 8, "Inversor:", 0, 0); pdf.cell(0, 8, str(inversores_str), 0, 1)
    pdf.cell(col1, 8, "Modulos:", 0, 0); pdf.cell(0, 8, str(modulos_str), 0, 1)
    pdf.ln(10)
    
    # Materials Table
    pdf.set_font("Arial", 'B', 12)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(140, 10, "Produto", 1, 0, 'L', True)
    pdf.cell(40, 10, "Quantidade", 1, 1, 'C', True)
    
    pdf.set_font("Arial", '', 11)
    for nome, info in materiais.items():
        # Handle current format {qtd: X, origem: Y} or just X
        qtd = info.get('qtd', info) if isinstance(info, dict) else info
        pdf.cell(140, 9, str(nome), 1)
        pdf.cell(40, 9, str(qtd), 1, 1, 'C')
        
    pdf_path = os.path.join(BASE_DIR, 'saidas', f"ROMANEIO_{cliente}_{datetime.now().strftime('%H%M%S')}.pdf")
    if not os.path.exists(os.path.join(BASE_DIR, 'saidas')): os.makedirs(os.path.join(BASE_DIR, 'saidas'))
    
    pdf.output(pdf_path)
    return send_file(pdf_path, as_attachment=True)

@app.route('/api/download/<path:filename>')
def download_arquivo(filename):
    """Download de um arquivo salvo anteriormente."""
    pasta = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'saidas')
    caminho = os.path.join(pasta, filename)
    caminho_real = os.path.realpath(caminho)
    pasta_real = os.path.realpath(pasta)
    if not caminho_real.startswith(pasta_real):
        return jsonify({'erro': 'Acesso negado'}), 403
    if not os.path.exists(caminho):
        return jsonify({'erro': 'Arquivo não encontrado'}), 404
    return send_file(caminho, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/banco', methods=['GET'])
def get_banco():
    return jsonify(load_db())

@app.route('/api/banco', methods=['POST'])
def save_banco():
    """Salva alterações no banco de dados."""
    try:
        novo = request.get_json()
        with open(os.path.join(BASE_DIR, 'database.json'), 'w', encoding='utf-8') as f:
            json.dump(novo, f, ensure_ascii=False, indent=2)
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/gerar_lote', methods=['POST'])
def gerar_lote():
    try:
        data_list = request.get_json() # Espera uma lista de projetos
        if not data_list or not isinstance(data_list, list):
            return jsonify({'erro': 'Lista de projetos inválida'}), 400

        db = load_db()
        wb = openpyxl.Workbook()
        # Remove a aba padrão criada pelo Workbook
        default_sheet = wb.active
        wb.remove(default_sheet)

        projetos_processados = []

        # RE-ABORDAGEM: Abrir o modelo e ir duplicando a aba interna dele para cada cliente
        wb_final = openpyxl.load_workbook(os.path.join(BASE_DIR, 'modelo.xlsx'))
        template_ws = wb_final.active
        template_ws.title = "TEMPLATE_BASE"

        for i, project in enumerate(data_list):
            cliente     = project.get('cliente', 'Sem Nome').strip()
            data_obra   = project.get('data', '').strip()
            equipe      = project.get('equipe', '').strip()
            cod_obra    = project.get('cod_obra', '').strip()
            modulos     = project.get('modulos', [])
            inversores  = project.get('inversores', [])
            extras      = project.get('extras', [])
            cidade      = project.get('cidade', '').strip()
            estrutura   = project.get('estrutura', '').strip()

            # Calcular
            quantidades = calcular_quantidades(modulos, inversores, cidade, db)

            # Duplicar
            new_ws = wb_final.copy_worksheet(template_ws)
            s_name = "".join(c for c in cliente if c.isalnum() or c in ' ' ).strip()[:30] or f"Projeto_{i+1}"
            if s_name in wb_final.sheetnames:
                s_name = f"{s_name[:27]}_{i+1}"
            new_ws.title = s_name

            # Preencher Cabeçalho
            def preencher_ajustado(ws, row, label, valor):
                orig = ws.cell(row=row, column=1).value or ""
                if label.strip() in orig:
                    ws.cell(row=row, column=1).value = f"{label}{valor}"
                else:
                    ws.cell(row=row, column=1).value = f"{label}: {valor}"

            inversores_str = ", ".join([f"{i.get('qtd', 1)}x {i.get('nome', '')}" for i in inversores])
            modulos_str = ", ".join([f"{m.get('qtd', 1)}x {m.get('potencia', '')}W" for m in modulos]) if modulos else "Nenhum"

            preencher_ajustado(new_ws, 2, "Data: ", data_obra)
            preencher_ajustado(new_ws, 3, "Cliente: ", cliente)
            preencher_ajustado(new_ws, 4, "Equipe: ", equipe)
            preencher_ajustado(new_ws, 5, "Inversor/Marca: ", inversores_str)
            preencher_ajustado(new_ws, 6, "Paineis (Qtde/Potencia): ", modulos_str)
            preencher_ajustado(new_ws, 7, "Cód obra: ", cod_obra)
            if estrutura:
                preencher_ajustado(new_ws, 8, "Estrutura: ", estrutura)

            # Criar map das linhas originais da planilha pela Coluna 1
            row_map = {}
            for r in range(9, 56):
                c = new_ws.cell(row=r, column=1).value
                if c is not None:
                    row_map[str(c).strip()] = r
                    
            # Dicionário Nome Normalizado -> Código
            nome_para_codigo = {normalize_name(p['nome']): str(p['codigo']) for p in db.get('produtos', [])}

            extra_row = 57 # "ITENS EXTRAS"
            for ext in extras:
                nome_ext = ext.get('nome', '').strip()
                qtd_ext = ext.get('qtd', '')
                cod = ext.get('codigo')
                if not nome_ext and not cod: continue
                
                nome_norm = normalize_name(nome_ext)
                # Tenta resolver código explícito ou implícito
                if cod and str(cod) in row_map:
                    new_ws.cell(row=row_map[str(cod)], column=3).value = qtd_ext
                elif nome_norm in nome_para_codigo and nome_para_codigo[nome_norm] in row_map:
                    new_ws.cell(row=row_map[nome_para_codigo[nome_norm]], column=3).value = qtd_ext
                else:
                    new_ws.cell(row=extra_row, column=2).value = nome_ext
                    new_ws.cell(row=extra_row, column=3).value = qtd_ext
                    extra_row += 1

            disj_ca_list = []
            disj_cc_list = []
            for inv in inversores:
                i_nome = inv.get('nome', '').strip()
                if i_nome in db.get('inversores', {}):
                    dca = db['inversores'][i_nome].get('disjuntor_ca', '')
                    dcc = db['inversores'][i_nome].get('disjuntor_cc', '')
                    if dca: disj_ca_list.append(f"{inv.get('qtd')}x {dca}" if inv.get('qtd', 1) > 1 else dca)
                    if dcc: disj_cc_list.append(f"{inv.get('qtd')}x {dcc}" if inv.get('qtd', 1) > 1 else dcc)
            
            if disj_ca_list:
                new_ws.cell(row=28, column=3).value = " / ".join(disj_ca_list)
            if disj_cc_list:
                new_ws.cell(row=29, column=3).value = " / ".join(disj_cc_list)

            for key, qtd in quantidades.items():
                nome = key
                cod = None
                if key.startswith('[') and '] ' in key:
                    cod = key.split('] ')[0][1:]
                    nome = key.split('] ')[1]
                    
                nome_norm = normalize_name(nome)
                if cod and str(cod) in row_map:
                    new_ws.cell(row=row_map[str(cod)], column=3).value = qtd
                elif nome_norm in nome_para_codigo and nome_para_codigo[nome_norm] in row_map:
                    new_ws.cell(row=row_map[nome_para_codigo[nome_norm]], column=3).value = qtd
                else:
                    # É um item não listado nativo, envia pra base extra
                    new_ws.cell(row=extra_row, column=2).value = nome
                    new_ws.cell(row=extra_row, column=3).value = qtd
                    extra_row += 1

            projetos_processados.append(cliente)

        # Remover a aba template original antes de salvar
        wb_final.remove(template_ws)

        # Salvar
        pasta_saidas = os.path.join(BASE_DIR, 'saidas')
        os.makedirs(pasta_saidas, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"LOTE_SSM_{timestamp}.xlsx"
        caminho_disco = os.path.join(pasta_saidas, filename)

        wb_final.save(caminho_disco)
        
        # Logar histórico do lote
        log_history(data_list)

        return send_file(
            caminho_disco,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    host = os.environ.get("HOST", "0.0.0.0")
    
    # Em produção, FLASK_DEBUG deve ser False
    is_development = os.environ.get("FLASK_ENV", "development") == "development"
    
    print("=" * 50)
    print("  SSM Solar — Sistema de Materiais v2.0")
    print(f"  Servidor rodando em: http://{host}:{port}")
    print("=" * 50)
    
    app.run(host=host, port=port, debug=is_development, use_reloader=False)
