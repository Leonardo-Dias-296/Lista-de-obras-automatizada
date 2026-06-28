import openpyxl

wb = openpyxl.load_workbook('modelo.xlsx')
ws = wb.active

print("=== MERGED CELLS ===")
for m in ws.merged_cells.ranges:
    print(f"  {m}  → min_row={m.min_row} min_col={m.min_col} max_row={m.max_row} max_col={m.max_col}")

print("\n=== HEADER ROWS 1-8 (all columns A-E) ===")
from openpyxl.cell.cell import MergedCell
for r in range(1, 9):
    for c in range(1, 6):
        cell = ws.cell(r, c)
        val  = cell.value
        tipo = "MERGED" if isinstance(cell, MergedCell) else "normal"
        if val is not None or tipo == "MERGED":
            print(f"  ({r},{c}) [{tipo}] = {repr(val)}")
