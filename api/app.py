from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import traceback
import json
import io
import os
import sys
import tempfile
from datetime import datetime

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_DIR = os.path.join(BASE_DIR, '..')

def _import_calcular():
    if BASE_DIR not in sys.path:
        sys.path.insert(0, BASE_DIR)
    from calcular import load_db, save_db, calcular_quantidades, normalize_name
    return load_db, save_db, calcular_quantidades, normalize_name

@app.route('/')
def index():
    return jsonify({"status": "ok", "message": "SSM Solar API v2.0"})

@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500

@app.route('/api/config')
def config():
    load_db, _, _, _ = _import_calcular()
    db = load_db()
    return jsonify(db)

@app.route('/api/gerar', methods=['POST'])
def gerar():
    try:
        import openpyxl
        load_db, _, calcular_quantidades, normalize_name = _import_calcular()

        data = request.get_json()

        cliente     = data.get('cliente', '').strip()
        data_obra   = data.get('data', '').strip()
        equipe      = data.get('equipe', '').strip()
        cod_obra    = data.get('cod_obra', '').strip()
        modulos     = data.get('modulos', [])
        inversores  = data.get('inversores', [])
        extras      = data.get('extras', [])
        cidade      = data.get('cidade', '').strip()
        estrutura   = data.get('estrutura', '').strip()

        db = load_db()
        quantidades = calcular_quantidades(modulos, inversores, cidade, db)

        modelo_path = os.path.join(REPO_DIR, 'public', 'modelo.xlsx')
        if not os.path.exists(modelo_path):
            modelo_path = os.path.join(REPO_DIR, 'modelo.xlsx')
        wb = openpyxl.load_workbook(modelo_path)
        ws = wb.active

        def preencher_ajustado(row, label, valor):
            orig = ws.cell(row=row, column=1).value or ""
            if label.strip() in orig:
                ws.cell(row=row, column=1).value = f"{label}{valor}"
            else:
                ws.cell(row=row, column=1).value = f"{label}: {valor}"

        inversores_str = ", ".join([f"{i.get('qtd', 1)}x {i.get('nome', '')}" for i in inversores])
        modulos_str = ", ".join([f"{m.get('qtd', 1)}x {m.get('potencia', '')}W" for m in modulos]) if modulos else "Nenhum"

        preencher_ajustado(2, "Data: ", data_obra)
        preencher_ajustado(3, "Cliente: ", cliente)
        preencher_ajustado(4, "Equipe: ", equipe)
        preencher_ajustado(5, "Inversor/Marca: ", inversores_str)
        preencher_ajustado(6, "Paineis (Qtde/Potencia): ", modulos_str)
        preencher_ajustado(7, "Cód obra: ", cod_obra)
        if estrutura:
            preencher_ajustado(8, "Estrutura: ", estrutura)

        row_map = {}
        for r in range(9, 56):
            c = ws.cell(row=r, column=1).value
            if c is not None:
                row_map[str(c).strip()] = r

        nome_para_codigo = {normalize_name(p['nome']): str(p['codigo']) for p in db.get('produtos', [])}

        extra_row = 57
        for ext in extras:
            nome_ext = ext.get('nome', '').strip()
            qtd_ext = ext.get('qtd', '')
            cod = ext.get('codigo')
            if not nome_ext and not cod:
                continue

            nome_norm = normalize_name(nome_ext)
            if cod and str(cod) in row_map:
                ws.cell(row=row_map[str(cod)], column=3).value = qtd_ext
            elif nome_norm in nome_para_codigo and nome_para_codigo[nome_norm] in row_map:
                ws.cell(row=row_map[nome_para_codigo[nome_norm]], column=3).value = qtd_ext
            else:
                ws.cell(row=extra_row, column=2).value = nome_ext
                ws.cell(row=extra_row, column=3).value = qtd_ext
                extra_row += 1

        disj_ca_lista = []
        disj_cc_lista = []
        for inv in inversores:
            i_nome = inv.get('nome', '').strip()
            if i_nome in db.get('inversores', {}):
                dca = db['inversores'][i_nome].get('disjuntor_ca', '')
                dcc = db['inversores'][i_nome].get('disjuntor_cc', '')
                if dca:
                    disj_ca_lista.append(f"{inv.get('qtd')}x {dca}" if inv.get('qtd', 1) > 1 else dca)
                if dcc:
                    disj_cc_lista.append(f"{inv.get('qtd')}x {dcc}" if inv.get('qtd', 1) > 1 else dcc)

        if disj_ca_lista:
            ws.cell(row=28, column=3).value = " / ".join(disj_ca_lista)
        if disj_cc_lista:
            ws.cell(row=29, column=3).value = " / ".join(disj_cc_lista)

        for key, qtd in quantidades.items():
            nome = key
            cod = None
            if key.startswith('[') and '] ' in key:
                cod = key.split('] ')[0][1:]
                nome = key.split('] ')[1]

            nome_norm = normalize_name(nome)
            if cod and str(cod) in row_map:
                ws.cell(row=row_map[str(cod)], column=3).value = qtd
            elif nome_norm in nome_para_codigo and nome_para_codigo[nome_norm] in row_map:
                ws.cell(row=row_map[nome_para_codigo[nome_norm]], column=3).value = qtd
            else:
                ws.cell(row=extra_row, column=2).value = nome
                ws.cell(row=extra_row, column=3).value = qtd
                extra_row += 1

        nome_seguro = "".join(c for c in cliente if c.isalnum() or c in ' _-').strip().replace(' ', '_')
        filename = f"SSM_{nome_seguro}_{data_obra}.xlsx"

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp.name)
        tmp.close()

        return send_file(
            tmp.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

@app.route('/api/preview', methods=['POST'])
def preview():
    try:
        load_db, _, calcular_quantidades, _ = _import_calcular()
        data       = request.get_json()
        modulos    = data.get('modulos', [])
        inversores = data.get('inversores', [])
        cidade     = data.get('cidade', '').strip()

        db = load_db()
        quantidades = calcular_quantidades(modulos, inversores, cidade, db)
        return jsonify(quantidades)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/arquivos')
def listar_arquivos():
    return jsonify([])

@app.route('/api/arquivos/<path:filename>', methods=['DELETE'])
def deletar_arquivo(filename):
    return jsonify({'status': 'ok'})

@app.route('/api/stats')
def get_stats():
    return jsonify({
        "total_orcamentos": 0,
        "total_modulos": 0,
        "top_inversores": {},
        "top_cidades": {}
    })

@app.route('/api/gerar_pdf', methods=['POST'])
def gerar_pdf():
    from fpdf import FPDF
    dados = request.json
    cliente = dados.get('cliente', 'Cliente')
    equipe = dados.get('equipe', '—')
    data_obra = dados.get('data', datetime.now().strftime("%Y-%m-%d"))
    modulos = dados.get('modulos', [])
    modulos_str = ", ".join([f"{m.get('qtd', 1)}x {m.get('potencia', '')}W" for m in modulos]) if modulos else "Nenhum"
    inversores = dados.get('inversores', [])
    inversores_str = ", ".join([f"{i.get('qtd', 1)}x {i.get('nome', '')}" for i in inversores]) if inversores else "—"
    materiais = dados.get('materiais', {})

    pdf = FPDF()
    pdf.add_page()

    pdf.set_fill_color(15, 23, 42)
    pdf.rect(0, 0, 210, 40, 'F')

    pdf.set_font("Arial", 'B', 20)
    pdf.set_text_color(251, 191, 36)
    pdf.cell(0, 15, "SSM SOLAR - ROMANEIO DE CARGA", 0, 1, 'C')

    pdf.set_font("Arial", '', 10)
    pdf.set_text_color(200, 200, 200)
    pdf.cell(0, 5, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
    pdf.ln(15)

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 10, "DADOS DA OBRA", "B", 1)
    pdf.set_font("Arial", '', 11)

    col1 = 40
    pdf.cell(col1, 8, "Cliente:", 0, 0)
    pdf.cell(0, 8, str(cliente), 0, 1)
    pdf.cell(col1, 8, "Data Obra:", 0, 0)
    pdf.cell(0, 8, str(data_obra), 0, 1)
    pdf.cell(col1, 8, "Equipe:", 0, 0)
    pdf.cell(0, 8, str(equipe), 0, 1)
    pdf.cell(col1, 8, "Inversor:", 0, 0)
    pdf.cell(0, 8, str(inversores_str), 0, 1)
    pdf.cell(col1, 8, "Modulos:", 0, 0)
    pdf.cell(0, 8, str(modulos_str), 0, 1)
    pdf.ln(10)

    pdf.set_font("Arial", 'B', 12)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(140, 10, "Produto", 1, 0, 'L', True)
    pdf.cell(40, 10, "Quantidade", 1, 1, 'C', True)

    pdf.set_font("Arial", '', 11)
    for nome, info in materiais.items():
        qtd = info.get('qtd', info) if isinstance(info, dict) else info
        pdf.cell(140, 9, str(nome), 1)
        pdf.cell(40, 9, str(qtd), 1, 1, 'C')

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf.output(tmp.name)
    tmp.close()

    return send_file(tmp.name, as_attachment=True,
                     download_name=f"ROMANEIO_{cliente}.pdf")

@app.route('/api/download/<path:filename>')
def download_arquivo(filename):
    return jsonify({'erro': 'Arquivo não encontrado'}), 404

@app.route('/api/banco', methods=['GET'])
def get_banco():
    load_db, _, _, _ = _import_calcular()
    return jsonify(load_db())

@app.route('/api/banco', methods=['POST'])
def save_banco():
    try:
        _, save_db, _, _ = _import_calcular()
        novo = request.get_json()
        save_db(novo)
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/gerar_lote', methods=['POST'])
def gerar_lote():
    try:
        import openpyxl
        load_db, _, calcular_quantidades, normalize_name = _import_calcular()

        data_list = request.get_json()
        if not data_list or not isinstance(data_list, list):
            return jsonify({'erro': 'Lista de projetos inválida'}), 400

        db = load_db()

        modelo_path = os.path.join(REPO_DIR, 'public', 'modelo.xlsx')
        if not os.path.exists(modelo_path):
            modelo_path = os.path.join(REPO_DIR, 'modelo.xlsx')
        wb_final = openpyxl.load_workbook(modelo_path)
        template_ws = wb_final.active
        template_ws.title = "TEMPLATE_BASE"

        for i, project in enumerate(data_list):
            cliente     = project.get('cliente', 'Sem Nome').strip()
            data_obra   = project.get('data', '').strip()
            equipe      = project.get('equipe', '').strip()
            cod_obra    = project.get('cod_obra', '').strip()
            modulos     = project.get('modulos', [])
            inversores  = project.get('inversores', [])
            extras      = project.get('extras', [])
            cidade      = project.get('cidade', '').strip()
            estrutura   = project.get('estrutura', '').strip()

            quantidades = calcular_quantidades(modulos, inversores, cidade, db)

            new_ws = wb_final.copy_worksheet(template_ws)
            s_name = "".join(c for c in cliente if c.isalnum() or c in ' ').strip()[:30] or f"Projeto_{i+1}"
            if s_name in wb_final.sheetnames:
                s_name = f"{s_name[:27]}_{i+1}"
            new_ws.title = s_name

            def preencher_ajustado(ws, row, label, valor):
                orig = ws.cell(row=row, column=1).value or ""
                if label.strip() in orig:
                    ws.cell(row=row, column=1).value = f"{label}{valor}"
                else:
                    ws.cell(row=row, column=1).value = f"{label}: {valor}"

            inversores_str = ", ".join([f"{i.get('qtd', 1)}x {i.get('nome', '')}" for i in inversores])
            modulos_str = ", ".join([f"{m.get('qtd', 1)}x {m.get('potencia', '')}W" for m in modulos]) if modulos else "Nenhum"

            preencher_ajustado(new_ws, 2, "Data: ", data_obra)
            preencher_ajustado(new_ws, 3, "Cliente: ", cliente)
            preencher_ajustado(new_ws, 4, "Equipe: ", equipe)
            preencher_ajustado(new_ws, 5, "Inversor/Marca: ", inversores_str)
            preencher_ajustado(new_ws, 6, "Paineis (Qtde/Potencia): ", modulos_str)
            preencher_ajustado(new_ws, 7, "Cód obra: ", cod_obra)
            if estrutura:
                preencher_ajustado(new_ws, 8, "Estrutura: ", estrutura)

            row_map = {}
            for r in range(9, 56):
                c = new_ws.cell(row=r, column=1).value
                if c is not None:
                    row_map[str(c).strip()] = r

            nome_para_codigo = {normalize_name(p['nome']): str(p['codigo']) for p in db.get('produtos', [])}

            extra_row = 57
            for ext in extras:
                nome_ext = ext.get('nome', '').strip()
                qtd_ext = ext.get('qtd', '')
                cod = ext.get('codigo')
                if not nome_ext and not cod:
                    continue

                nome_norm = normalize_name(nome_ext)
                if cod and str(cod) in row_map:
                    new_ws.cell(row=row_map[str(cod)], column=3).value = qtd_ext
                elif nome_norm in nome_para_codigo and nome_para_codigo[nome_norm] in row_map:
                    new_ws.cell(row=row_map[nome_para_codigo[nome_norm]], column=3).value = qtd_ext
                else:
                    new_ws.cell(row=extra_row, column=2).value = nome_ext
                    new_ws.cell(row=extra_row, column=3).value = qtd_ext
                    extra_row += 1

            disj_ca_list = []
            disj_cc_list = []
            for inv in inversores:
                i_nome = inv.get('nome', '').strip()
                if i_nome in db.get('inversores', {}):
                    dca = db['inversores'][i_nome].get('disjuntor_ca', '')
                    dcc = db['inversores'][i_nome].get('disjuntor_cc', '')
                    if dca:
                        disj_ca_list.append(f"{inv.get('qtd')}x {dca}" if inv.get('qtd', 1) > 1 else dca)
                    if dcc:
                        disj_cc_list.append(f"{inv.get('qtd')}x {dcc}" if inv.get('qtd', 1) > 1 else dcc)

            if disj_ca_list:
                new_ws.cell(row=28, column=3).value = " / ".join(disj_ca_list)
            if disj_cc_list:
                new_ws.cell(row=29, column=3).value = " / ".join(disj_cc_list)

            for key, qtd in quantidades.items():
                nome = key
                cod = None
                if key.startswith('[') and '] ' in key:
                    cod = key.split('] ')[0][1:]
                    nome = key.split('] ')[1]

                nome_norm = normalize_name(nome)
                if cod and str(cod) in row_map:
                    new_ws.cell(row=row_map[str(cod)], column=3).value = qtd
                elif nome_norm in nome_para_codigo and nome_para_codigo[nome_norm] in row_map:
                    new_ws.cell(row=row_map[nome_para_codigo[nome_norm]], column=3).value = qtd
                else:
                    new_ws.cell(row=extra_row, column=2).value = nome
                    new_ws.cell(row=extra_row, column=3).value = qtd
                    extra_row += 1

        wb_final.remove(template_ws)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"LOTE_SSM_{timestamp}.xlsx"

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb_final.save(tmp.name)
        tmp.close()

        return send_file(
            tmp.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

if __name__ == '__main__':
    app.run(port=5000, debug=True)
